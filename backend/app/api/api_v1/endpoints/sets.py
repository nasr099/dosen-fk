from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, Request
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import List, Optional
from openpyxl_image_loader import SheetImageLoader
from openpyxl.utils import get_column_letter
from io import BytesIO
from app.api.deps import get_current_admin_user, get_current_staff_user
from app.db.base import get_db
from app.core.config import settings
from app.db.models import (
    QuestionSet as QuestionSetModel,
    Category as CategoryModel,
    Question as QuestionModel,
    ExamSession as ExamSessionModel,
    ExamAnswer as ExamAnswerModel,
    EssayGrade as EssayGradeModel,
    Reading as ReadingModel,
)
from app.schemas.set import (
    QuestionSet as QuestionSetSchema,
    QuestionSetCreate,
    QuestionSetUpdate,
    QuestionSetWithQuestionsCreate,
    QuestionSetWithQuestionsUpdate,
)

router = APIRouter()

# --- Helpers: robust embedded image extraction from XLSX ---
def _normalize_path(base: str, target: str) -> str:
    import posixpath
    # If base is a .rels inside a _rels folder, resolve relative to the source part's directory
    # Example: base = 'xl/worksheets/_rels/sheet1.xml.rels', target = '../drawings/drawing1.xml'
    # Source part is 'xl/worksheets/sheet1.xml' -> base dir 'xl/worksheets'
    base_dir = posixpath.dirname(base)
    if base.endswith('.rels') and '/_rels/' in base:
        source_part = base.replace('/_rels/', '/').replace('.rels', '')
        base_dir = posixpath.dirname(source_part)
    return posixpath.normpath(posixpath.join(base_dir, target))

def extract_embedded_images_by_cell(xlsx_bytes: bytes, sheet_index_1_based: int):
    """
    Return list of { 'row': int, 'col': int, 'data': bytes, 'content_type': str }
    for images anchored on the given sheet index (1-based).
    Uses DrawingML anchors from xl/drawings/drawing*.xml and rels to xl/media/*.
    """
    from io import BytesIO
    import zipfile
    import xml.etree.ElementTree as ET
    buf = BytesIO(xlsx_bytes)
    zf = zipfile.ZipFile(buf)
    ns = {
        'xdr': 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing',
        'a': 'http://schemas.openxmlformats.org/drawingml/2006/main',
        'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
    }
    sheet_xml = f"xl/worksheets/sheet{sheet_index_1_based}.xml"
    sheet_rels = f"xl/worksheets/_rels/sheet{sheet_index_1_based}.xml.rels"
    try:
        rels_root = ET.fromstring(zf.read(sheet_rels)) if sheet_rels in zf.namelist() else None
    except Exception:
        rels_root = None
    drawing_target = None
    if rels_root is not None:
        for rel in rels_root.findall('.//Relationship'):
            rtype = rel.attrib.get('Type', '')
            if rtype.endswith('/drawing'):
                drawing_target = rel.attrib.get('Target')
                if drawing_target:
                    drawing_target = _normalize_path(sheet_rels, drawing_target)
                break
    drawings_to_scan = []
    if drawing_target and drawing_target in zf.namelist():
        drawings_to_scan.append(drawing_target)
    else:
        # Fallback: scan all drawings
        drawings_to_scan = [n for n in zf.namelist() if n.startswith('xl/drawings/drawing') and n.endswith('.xml')]
        if not drawings_to_scan:
            return []

    out = []
    for dpath in drawings_to_scan:
        # map rId -> media path for this drawing
        draw_rels_path = _normalize_path(dpath, dpath.replace('drawings/', 'drawings/_rels/') + '.rels')
        media_map = {}
        try:
            if draw_rels_path in zf.namelist():
                dr = ET.fromstring(zf.read(draw_rels_path))
                for rel in dr.findall('.//Relationship'):
                    rid = rel.attrib.get('Id')
                    tgt = rel.attrib.get('Target')
                    if rid and tgt:
                        media_map[rid] = _normalize_path(draw_rels_path, tgt)
        except Exception:
            media_map = {}
        # parse anchors in this drawing
        try:
            root = ET.fromstring(zf.read(dpath))
        except Exception:
            continue
        anchors = []
        anchors += root.findall('.//xdr:twoCellAnchor', ns)
        anchors += root.findall('.//xdr:oneCellAnchor', ns)
        for a in anchors:
            frm = a.find('xdr:from', ns)
            if frm is None:
                continue
            try:
                fr = int(frm.find('xdr:row', ns).text) + 1
                fc = int(frm.find('xdr:col', ns).text) + 1
            except Exception:
                continue
            # try get 'to' for twoCellAnchor
            tr = fr
            tc = fc
            to = a.find('xdr:to', ns)
            if to is not None:
                try:
                    tr = int(to.find('xdr:row', ns).text) + 1
                    tc = int(to.find('xdr:col', ns).text) + 1
                except Exception:
                    tr, tc = fr, fc
            blip = a.find('.//a:blip', ns)
            rid = blip.attrib.get('{%s}embed' % ns['r']) if blip is not None else None
            media_path = media_map.get(rid)
            if not media_path:
                continue
            try:
                data = zf.read(media_path)
            except Exception:
                continue
            ext = media_path.split('.')[-1].lower()
            ctype = {
                'png':'image/png','jpg':'image/jpeg','jpeg':'image/jpeg','gif':'image/gif','webp':'image/webp','bmp':'image/bmp'
            }.get(ext, 'application/octet-stream')
            out.append({ 'row': fr, 'col': fc, 'from_row': fr, 'from_col': fc, 'to_row': tr, 'to_col': tc, 'data': data, 'content_type': ctype })
    return out

@router.get("/import-template.xlsx")
def download_import_template():
    try:
        from openpyxl import Workbook
    except Exception:
        raise HTTPException(status_code=500, detail="openpyxl is not installed on the server")

    wb = Workbook()
    ws = wb.active
    ws.title = "questions"
    headers = [
        'type','question_text','question_img',
        'option_a_text','option_a_img',
        'option_b_text','option_b_img',
        'option_c_text','option_c_img',
        'option_d_text','option_d_img',
        'option_e_text','option_e_img',
        'correct_answer','explanation',
        # Optional reading linkage
        'reading_id','reading_title','reading_content',
    ]
    ws.append(headers)
    # Example rows
    ws.append(['mcq','What is 2+2?','', '3','', '4','', '5','', '6','', '','', 'B','Basic arithmetic'])
    ws.append(['multi','Select primes','', '2','', '3','', '4','', '5','', '6','', 'A,B,D',''])
    ws.append(['essay','Explain the cardiac cycle.','', '','','','','','','','','','', ''])

    from io import BytesIO
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers={
        'Content-Disposition': 'attachment; filename="question_import_template.xlsx"'
    })

 

@router.get("/", response_model=List[QuestionSetSchema])
def list_sets(
    db: Session = Depends(get_db),
    category_id: Optional[int] = Query(None)
):
    q = db.query(QuestionSetModel).filter(QuestionSetModel.is_active == True)
    if category_id is not None:
        q = q.filter(QuestionSetModel.category_id == category_id)
    return q.order_by(QuestionSetModel.created_at.desc()).all()

@router.get("/summary")
def list_sets_summary(
    db: Session = Depends(get_db),
    category_id: Optional[int] = Query(None)
):
    """
    Return sets with aggregated question counts without fetching all questions.
    Response: [{ id, category_id, title, description, time_limit_minutes, access_level, created_at, updated_at, count }]
    """
    from sqlalchemy import func
    from sqlalchemy.orm import aliased
    qs = QuestionSetModel
    q = QuestionModel
    base = db.query(
        qs.id,
        qs.category_id,
        qs.title,
        qs.description,
        qs.time_limit_minutes,
        qs.is_active,
        qs.access_level,
        qs.created_at,
        qs.updated_at,
        func.count(q.id).label('count'),
    ).outerjoin(q, q.question_set_id == qs.id)
    base = base.filter(qs.is_active == True)
    if category_id is not None:
        base = base.filter(qs.category_id == category_id)
    rows = base.group_by(
        qs.id, qs.category_id, qs.title, qs.description, qs.time_limit_minutes, qs.is_active, qs.access_level, qs.created_at, qs.updated_at
    ).order_by(qs.created_at.desc()).all()
    return [
        {
            'id': r.id,
            'category_id': r.category_id,
            'title': r.title,
            'description': r.description,
            'time_limit_minutes': r.time_limit_minutes,
            'is_active': r.is_active,
            'access_level': r.access_level,
            'created_at': r.created_at,
            'updated_at': r.updated_at,
            'count': int(r.count or 0),
        }
        for r in rows
    ]

@router.get("/{set_id}", response_model=QuestionSetSchema)
def get_set(set_id: int, db: Session = Depends(get_db)):
    s = db.query(QuestionSetModel).filter(QuestionSetModel.id == set_id).first()
    if not s:
        raise HTTPException(status_code=404, detail="Set not found")
    return s

@router.put("/{set_id}/with-questions", response_model=QuestionSetSchema)
def update_set_with_questions(
    set_id: int,
    payload: QuestionSetWithQuestionsUpdate,
    db: Session = Depends(get_db),
    admin=Depends(get_current_staff_user),
):
    s = db.query(QuestionSetModel).filter(QuestionSetModel.id == set_id).first()
    if not s:
        raise HTTPException(status_code=404, detail="Set not found")

    # update set fields
    for k, v in payload.model_dump(exclude_unset=True, exclude={"questions"}).items():
        setattr(s, k, v)
    db.commit()
    db.refresh(s)

    # replace questions belonging to this set
    # First, delete essay_grades -> exam_answers that reference existing questions in this set (respect FK)
    q_sub = db.query(QuestionModel.id).filter(QuestionModel.question_set_id == set_id).subquery()
    ans_sub = db.query(ExamAnswerModel.id).filter(ExamAnswerModel.question_id.in_(q_sub)).subquery()
    db.query(EssayGradeModel).filter(EssayGradeModel.exam_answer_id.in_(ans_sub)).delete(synchronize_session=False)
    db.commit()
    # Then delete the exam_answers
    db.query(ExamAnswerModel).filter(ExamAnswerModel.question_id.in_(q_sub)).delete(synchronize_session=False)
    db.commit()
    # Then remove questions in this set
    db.query(QuestionModel).filter(QuestionModel.question_set_id == set_id).delete(synchronize_session=False)
    db.commit()
    for q in payload.questions:
        db.add(QuestionModel(
            category_id=s.category_id,
            question_set_id=set_id,
            question_text=q.question_text,
            question_type=getattr(q, 'question_type', 'mcq') or 'mcq',
            reading_id=getattr(q, 'reading_id', None),
            option_a=q.option_a,
            option_b=q.option_b,
            option_c=q.option_c,
            option_d=q.option_d,
            option_e=q.option_e,
            correct_answer=q.correct_answer,
            explanation=q.explanation,
            is_featured=q.is_featured,
            difficulty_level=q.difficulty_level or 'medium',
        ))
    db.commit()
    db.refresh(s)
    return s

@router.post("/{set_id}/import-xlsx")
def import_questions_from_xlsx(
    set_id: int,
    file: UploadFile = File(...),
    request: Request = None,
    db: Session = Depends(get_db),
    admin=Depends(get_current_staff_user)
):
    """
    Import questions from an .xlsx file. Expected columns (case-insensitive headers):
    - type: mcq | multi | essay
    - question_text, question_img (optional)
    - option_a_text, option_a_img (optional)
    - option_b_text, option_b_img (optional)
    - option_c_text, option_c_img (optional)
    - option_d_text, option_d_img (optional)
    - option_e_text, option_e_img (optional)
    - correct_answer (single letter A-E for mcq, comma-separated letters for multi, blank for essay)
    - explanation (optional)
    """
    # validate set
    s = db.query(QuestionSetModel).filter(QuestionSetModel.id == set_id).first()
    if not s:
        raise HTTPException(status_code=404, detail="Set not found")

    try:
        from openpyxl import load_workbook
    except Exception:
        raise HTTPException(status_code=500, detail="openpyxl is not installed on the server")

    # read workbook
    try:
        from io import BytesIO
        # Read uploaded content into bytes buffer (some SpooledTemporaryFile may not be seekable as expected)
        content = file.file.read()
        if not content:
            raise ValueError('empty file')
        buf = BytesIO(content)
        wb = load_workbook(filename=buf, data_only=True)
        ws = wb.active
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to read xlsx: {e}")

    # map headers
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="Empty worksheet")
    headers = [str(h or '').strip().lower() for h in rows[0]]
    col_index = { h: i for i, h in enumerate(headers) if h }

    # Image loader to read images by cell address
    try:
        image_loader = SheetImageLoader(ws)
    except Exception:
        image_loader = None

    def get(row, name, default=''):
        i = col_index.get(name)
        if i is None or i >= len(row):
            return default
        v = row[i]
        return '' if v is None else str(v).strip()

    def to_rich(text, img):
        t = (text or '').strip()
        i = (img or '').strip()
        if not i:
            return t
        return f'{{"text":"{t}","img":"{i}"}}'

    # Uploader settings
    DO_BUCKET = settings.DO_SPACES_BUCKET
    DO_REGION = settings.DO_SPACES_REGION or 'sgp1'
    DO_ENDPOINT = settings.DO_SPACES_ENDPOINT or f'https://{DO_REGION}.digitaloceanspaces.com'
    DO_KEY = settings.DO_SPACES_KEY
    DO_SECRET = settings.DO_SPACES_SECRET
    PUBLIC_CDN_BASE = settings.PUBLIC_CDN_BASE
    spaces_ok = bool(DO_BUCKET and DO_KEY and DO_SECRET and DO_ENDPOINT)

    def upload_bytes(data: bytes, content_type: str) -> str:
        try:
            if spaces_ok:
                import boto3
                from botocore.client import Config
                from uuid import uuid4
                ext = {
                    'image/jpeg':'.jpg', 'image/jpg':'.jpg', 'image/png':'.png', 'image/gif':'.gif', 'image/webp':'.webp',
                    'image/svg+xml':'.svg'
                }.get((content_type or '').split(';')[0].lower(), '')
                key = f"uploads/import-{uuid4().hex}{ext}"
                s3 = boto3.client('s3', region_name=DO_REGION, endpoint_url=DO_ENDPOINT,
                                  aws_access_key_id=DO_KEY, aws_secret_access_key=DO_SECRET,
                                  config=Config(signature_version='s3v4'))
                s3.put_object(Bucket=DO_BUCKET, Key=key, Body=data, ContentType=content_type or 'application/octet-stream', ACL='public-read', CacheControl='public, max-age=31536000, immutable')
                public_base = PUBLIC_CDN_BASE or f"https://{DO_BUCKET}.{DO_REGION}.digitaloceanspaces.com"
                return f"{public_base}/{key}"
        except Exception:
            pass
        # fallback to local
        import os
        from uuid import uuid4
        UPLOAD_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), '../../../uploads'))
        os.makedirs(UPLOAD_DIR, exist_ok=True)
        name = f"import-{uuid4().hex}.bin"
        path = os.path.join(UPLOAD_DIR, name)
        with open(path, 'wb') as f:
            f.write(data)
        base = str(request.base_url).rstrip('/') if request else ''
        return f"{base}/uploads/{name}" if base else f"/uploads/{name}"

    created = 0
    errors = []

    # start at row 2 (index 1)
    for idx, row in enumerate(rows[1:], start=2):
        qtype = (get(row, 'type') or 'mcq').lower().strip()
        if qtype not in ('mcq','multi','essay'):
            errors.append(f"Row {idx}: invalid type '{qtype}'")
            continue
        qtext = get(row, 'question_text')
        qimg_raw = get(row, 'question_img')
        r_idx = idx
        qimg = ''
        c_qimg = col_index.get('question_img')
        if image_loader and c_qimg is not None:
            from openpyxl.utils import get_column_letter
            addr = f"{get_column_letter(c_qimg+1)}{r_idx}"
            try:
                if image_loader.image_in(addr):
                    pil_img = image_loader.get(addr)
                    from io import BytesIO
                    buf = BytesIO()
                    fmt = (getattr(pil_img, 'format', None) or 'PNG').upper()
                    pil_img.save(buf, format=fmt)
                    data = buf.getvalue()
                    mime = {
                        'PNG':'image/png','JPG':'image/jpeg','JPEG':'image/jpeg','GIF':'image/gif','WEBP':'image/webp','BMP':'image/bmp'
                    }.get(fmt, 'image/png')
                    qimg = upload_bytes(data, mime)
            except Exception:
                pass
        if not qimg:
            # ignore Excel error tokens like #VALUE!
            qimg = '' if str(qimg_raw).startswith('#') else str(qimg_raw)
        if not qtext and not qimg:
            errors.append(f"Row {idx}: question_text or question_img is required")
            continue

        opts = {}
        letters = ['a','b','c','d','e']
        for l in letters:
            ot = get(row, f'option_{l}_text')
            oi_raw = get(row, f'option_{l}_img')
            oi = ''
            c_img = col_index.get(f'option_{l}_img')
            if image_loader and c_img is not None:
                from openpyxl.utils import get_column_letter
                addr2 = f"{get_column_letter(c_img+1)}{r_idx}"
                try:
                    if image_loader.image_in(addr2):
                        pil_img2 = image_loader.get(addr2)
                        from io import BytesIO
                        buf2 = BytesIO()
                        fmt2 = (getattr(pil_img2, 'format', None) or 'PNG').upper()
                        pil_img2.save(buf2, format=fmt2)
                        data2 = buf2.getvalue()
                        mime2 = {
                            'PNG':'image/png','JPG':'image/jpeg','JPEG':'image/jpeg','GIF':'image/gif','WEBP':'image/webp','BMP':'image/bmp'
                        }.get(fmt2, 'image/png')
                        oi = upload_bytes(data2, mime2)
                except Exception:
                    pass
            if not oi:
                oi = '' if str(oi_raw).startswith('#') else str(oi_raw)
            opts[l] = to_rich(ot, oi)

        correct_answer = get(row, 'correct_answer').upper().replace(' ', '')
        if qtype == 'mcq':
            if correct_answer not in ['A','B','C','D','E']:
                errors.append(f"Row {idx}: correct_answer must be one of A,B,C,D,E for mcq")
                continue
        elif qtype == 'multi':
            # Allow A,B or A;B formats
            sep = ',' if ',' in correct_answer else (';' if ';' in correct_answer else ',')
            parts = [p for p in [x.strip() for x in correct_answer.split(sep)] if p]
            if not parts:
                errors.append(f"Row {idx}: correct_answer must include at least one letter for multi")
                continue
            if any(p not in ['A','B','C','D','E'] for p in parts):
                errors.append(f"Row {idx}: correct_answer must contain letters A-E for multi")
                continue
            correct_answer = ','.join(sorted(set(parts)))
        else:
            # essay: blank correct
            correct_answer = ''

        # Resolve reading
        reading_id_val = None
        rid_raw = get(row, 'reading_id')
        rtitle = get(row, 'reading_title')
        rcontent = get(row, 'reading_content')
        try:
            if rid_raw:
                reading_id_val = int(float(rid_raw))
        except Exception:
            reading_id_val = None
        if not reading_id_val and (rtitle and rcontent):
            # find or create by exact title
            rd = db.query(ReadingModel).filter(ReadingModel.title == rtitle).first()
            if not rd:
                rd = ReadingModel(title=rtitle, content_html=rcontent, category_id=s.category_id)
                db.add(rd)
                db.commit()
                db.refresh(rd)
            reading_id_val = rd.id

        q = QuestionModel(
            category_id=s.category_id,
            question_set_id=set_id,
            question_text=to_rich(qtext, qimg),
            question_type=qtype,
            reading_id=reading_id_val,
            option_a=opts['a'], option_b=opts['b'], option_c=opts['c'], option_d=opts['d'], option_e=opts['e'],
            correct_answer=correct_answer,
            explanation=get(row, 'explanation'),
            is_featured=False,
            difficulty_level='medium',
        )
        db.add(q)
        created += 1

    db.commit()

    return {
        "created": created,
        "errors": errors,
    }

@router.post("/{set_id}/import-xlsx/preview")
def preview_import_questions_from_xlsx(
    set_id: int,
    file: UploadFile = File(...),
    request: Request = None,
    db: Session = Depends(get_db),
    admin=Depends(get_current_staff_user)
):
    # Reuse the same parsing logic as import, but do not write to DB; return parsed rows
    s = db.query(QuestionSetModel).filter(QuestionSetModel.id == set_id).first()
    if not s:
        raise HTTPException(status_code=404, detail="Set not found")
    try:
        from openpyxl import load_workbook
        from io import BytesIO
    except Exception:
        raise HTTPException(status_code=500, detail="openpyxl is not installed on the server")

    try:
        content = file.file.read()
        if not content:
            raise ValueError('empty file')
        buf = BytesIO(content)
        wb = load_workbook(filename=buf, data_only=True)
        ws = wb.active
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to read xlsx: {e}")

    # Headers and helper functions must mirror the import route
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="Empty worksheet")
    headers = [str(h or '').strip().lower() for h in rows[0]]
    col_index = { h: i for i, h in enumerate(headers) if h }

    # Image loader for reading images by cell address
    try:
        image_loader = SheetImageLoader(ws)
    except Exception:
        image_loader = None

    def get(row, name, default=''):
        i = col_index.get(name)
        if i is None or i >= len(row):
            return default
        v = row[i]
        return '' if v is None else str(v).strip()

    def detect_content_type(b: bytes) -> str:
        if not b:
            return 'application/octet-stream'
        if b.startswith(b"\x89PNG"): return 'image/png'
        if b.startswith(b"\xff\xd8"): return 'image/jpeg'
        if b.startswith(b"GIF8"): return 'image/gif'
        if b.startswith(b"RIFF") and b[8:12] == b"WEBP": return 'image/webp'
        return 'application/octet-stream'

    # For preview, upload embedded images to local /uploads or CDN as in import route
    DO_BUCKET = settings.DO_SPACES_BUCKET
    DO_REGION = settings.DO_SPACES_REGION or 'sgp1'
    DO_ENDPOINT = settings.DO_SPACES_ENDPOINT or f'https://{DO_REGION}.digitaloceanspaces.com'
    DO_KEY = settings.DO_SPACES_KEY
    DO_SECRET = settings.DO_SPACES_SECRET
    PUBLIC_CDN_BASE = settings.PUBLIC_CDN_BASE
    spaces_ok = bool(DO_BUCKET and DO_KEY and DO_SECRET and DO_ENDPOINT)

    def upload_bytes(data: bytes, content_type: str):
        """Return (url, diag). diag is None on success, otherwise a short string.
        """
        try:
            if spaces_ok:
                import boto3
                from botocore.client import Config
                from uuid import uuid4
                ext = {
                    'image/jpeg':'.jpg', 'image/jpg':'.jpg', 'image/png':'.png', 'image/gif':'.gif', 'image/webp':'.webp',
                    'image/svg+xml':'.svg'
                }.get((content_type or '').split(';')[0].lower(), '')
                key = f"uploads/preview-{uuid4().hex}{ext}"
                s3 = boto3.client('s3', region_name=DO_REGION, endpoint_url=DO_ENDPOINT,
                                  aws_access_key_id=DO_KEY, aws_secret_access_key=DO_SECRET,
                                  config=Config(signature_version='s3v4'))
                s3.put_object(Bucket=DO_BUCKET, Key=key, Body=data, ContentType=content_type or 'application/octet-stream', ACL='public-read', CacheControl='public, max-age=31536000, immutable')
                public_base = PUBLIC_CDN_BASE or f"https://{DO_BUCKET}.{DO_REGION}.digitaloceanspaces.com"
                return f"{public_base}/{key}", None
        except Exception as e:
            return '', f"spaces_upload_error: {e}"
        # fallback to local
        import os
        from uuid import uuid4
        UPLOAD_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), '../../../uploads'))
        os.makedirs(UPLOAD_DIR, exist_ok=True)
        name = f"preview-{uuid4().hex}.bin"
        path = os.path.join(UPLOAD_DIR, name)
        with open(path, 'wb') as f:
            f.write(data)
        base = str(request.base_url).rstrip('/') if request else ''
        url = f"{base}/uploads/{name}" if base else f"/uploads/{name}"
        if not spaces_ok:
            return url, "spaces_not_configured_fallback_local"
        return url, None

    preview_rows = []
    errors = []
    diagnostics = { 'spaces_ok': spaces_ok, 'notes': [], 'headers': headers, 'detected_cells': [], 'sheet_title': ws.title if hasattr(ws, 'title') else 'unknown' }
    letters = ['a','b','c','d','e']
    for idx, row in enumerate(rows[1:], start=2):
        r_idx = idx
        obj = { 'row': idx }
        obj['type'] = (get(row, 'type') or 'mcq').lower().strip()
        obj['question_text'] = get(row, 'question_text')
        # question image via image-loader
        c_qimg = col_index.get('question_img')
        url = ''
        if image_loader and c_qimg is not None:
            addr = f"{get_column_letter(c_qimg+1)}{r_idx}"
            try:
                if image_loader.image_in(addr):
                    pil_img = image_loader.get(addr)
                    buf = BytesIO()
                    fmt = (getattr(pil_img, 'format', None) or 'PNG').upper()
                    pil_img.save(buf, format=fmt)
                    data = buf.getvalue()
                    mime = {
                        'PNG':'image/png','JPG':'image/jpeg','JPEG':'image/jpeg','GIF':'image/gif','WEBP':'image/webp','BMP':'image/bmp'
                    }.get(fmt, 'image/png')
                    url, d = upload_bytes(data, mime)
                    diagnostics['detected_cells'].append({'cell': addr})
                    if d:
                        diagnostics['notes'].append({ 'row': r_idx, 'field': 'question_img', 'reason': d })
            except Exception:
                pass
        if not url:
            raw = get(row, 'question_img')
            url = '' if str(raw).startswith('#') else str(raw)
            if not url:
                diagnostics['notes'].append({ 'row': r_idx, 'field': 'question_img', 'reason': 'no_embedded_image_and_no_url_in_cell' })
        obj['question_img'] = url
        # options
        for l in letters:
            obj[f'option_{l}_text'] = get(row, f'option_{l}_text')
            c_img = col_index.get(f'option_{l}_img')
            val = ''
            if image_loader and c_img is not None:
                addr2 = f"{get_column_letter(c_img+1)}{r_idx}"
                try:
                    if image_loader.image_in(addr2):
                        pil_img2 = image_loader.get(addr2)
                        buf2 = BytesIO()
                        fmt2 = (getattr(pil_img2, 'format', None) or 'PNG').upper()
                        pil_img2.save(buf2, format=fmt2)
                        data2 = buf2.getvalue()
                        mime2 = {
                            'PNG':'image/png','JPG':'image/jpeg','JPEG':'image/jpeg','GIF':'image/gif','WEBP':'image/webp','BMP':'image/bmp'
                        }.get(fmt2, 'image/png')
                        url2, d2 = upload_bytes(data2, mime2)
                        diagnostics['detected_cells'].append({'cell': addr2})
                        if d2:
                            diagnostics['notes'].append({ 'row': r_idx, 'field': f'option_{l}_img', 'reason': d2 })
                        obj[f'option_{l}_img'] = url2
                        continue
                except Exception:
                    pass
            raw = get(row, f'option_{l}_img')
            val = '' if str(raw).startswith('#') else str(raw)
            if not val:
                diagnostics['notes'].append({ 'row': r_idx, 'field': f'option_{l}_img', 'reason': 'no_embedded_image_and_no_url_in_cell' })
            obj[f'option_{l}_img'] = val
        obj['correct_answer'] = get(row, 'correct_answer')
        obj['explanation'] = get(row, 'explanation')
        preview_rows.append(obj)

    return { 'rows': preview_rows, 'errors': errors, 'diagnostics': diagnostics }

 

@router.post("/with-questions", response_model=QuestionSetSchema)
def create_set_with_questions(
    payload: QuestionSetWithQuestionsCreate,
    db: Session = Depends(get_db),
    admin=Depends(get_current_staff_user)
):
    # validate category
    cat = db.query(CategoryModel).filter(CategoryModel.id == payload.category_id).first()
    if not cat:
        raise HTTPException(status_code=400, detail="Invalid category")
    s = QuestionSetModel(
        category_id=payload.category_id,
        title=payload.title,
        description=payload.description,
        time_limit_minutes=payload.time_limit_minutes,
        is_active=payload.is_active,
        access_level=getattr(payload, 'access_level', 'free') or 'free',
    )
    db.add(s)
    db.commit()
    db.refresh(s)

    # create questions under this set
    for q in payload.questions:
        db.add(QuestionModel(
            category_id=payload.category_id,
            question_set_id=s.id,
            question_text=q.question_text,
            question_type=getattr(q, 'question_type', 'mcq') or 'mcq',
            reading_id=getattr(q, 'reading_id', None),
            option_a=q.option_a,
            option_b=q.option_b,
            option_c=q.option_c,
            option_d=q.option_d,
            option_e=q.option_e,
            correct_answer=q.correct_answer,
            explanation=q.explanation,
            is_featured=q.is_featured,
            difficulty_level=q.difficulty_level or 'medium',
        ))
    db.commit()
    db.refresh(s)
    return s

@router.post("/", response_model=QuestionSetSchema)
def create_set(
    payload: QuestionSetCreate,
    db: Session = Depends(get_db),
    admin=Depends(get_current_staff_user)
):
    # validate category
    cat = db.query(CategoryModel).filter(CategoryModel.id == payload.category_id).first()
    if not cat:
        raise HTTPException(status_code=400, detail="Invalid category")
    s = QuestionSetModel(**payload.model_dump())
    db.add(s)
    db.commit()
    db.refresh(s)
    return s

@router.put("/{set_id}", response_model=QuestionSetSchema)
def update_set(
    set_id: int,
    payload: QuestionSetUpdate,
    db: Session = Depends(get_db),
    admin=Depends(get_current_staff_user)
):
    s = db.query(QuestionSetModel).filter(QuestionSetModel.id == set_id).first()
    if not s:
        raise HTTPException(status_code=404, detail="Set not found")
    for k, v in payload.model_dump(exclude_unset=True).items():
        setattr(s, k, v)
    db.commit()
    db.refresh(s)
    return s

@router.delete("/{set_id}")
def delete_set(set_id: int, db: Session = Depends(get_db), admin=Depends(get_current_staff_user)):
    s = db.query(QuestionSetModel).filter(QuestionSetModel.id == set_id).first()
    if not s:
        raise HTTPException(status_code=404, detail="Set not found")
    # 1) delete essay_grades then exam_answers that reference questions in this set
    q_sub = db.query(QuestionModel.id).filter(QuestionModel.question_set_id == set_id).subquery()
    ans_sub = db.query(ExamAnswerModel.id).filter(ExamAnswerModel.question_id.in_(q_sub)).subquery()
    db.query(EssayGradeModel).filter(EssayGradeModel.exam_answer_id.in_(ans_sub)).delete(synchronize_session=False)
    db.commit()
    db.query(ExamAnswerModel).filter(ExamAnswerModel.question_id.in_(q_sub)).delete(synchronize_session=False)
    db.commit()
    # 2) remove questions under this set to satisfy FK
    db.query(QuestionModel).filter(QuestionModel.question_set_id == set_id).delete(synchronize_session=False)
    db.commit()
    # 3) null-out exam sessions referencing this set (keep history intact)
    db.query(ExamSessionModel).filter(ExamSessionModel.question_set_id == set_id).update({ExamSessionModel.question_set_id: None})
    db.commit()
    # 4) delete the set itself
    db.delete(s)
    db.commit()
    return {"message": "Set deleted"}
