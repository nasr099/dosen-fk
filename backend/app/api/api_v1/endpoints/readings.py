from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import List, Optional
from io import BytesIO
from app.api.deps import get_current_staff_user, get_current_admin_user
from app.db.base import get_db
from app.db.models import Reading as ReadingModel
from app.schemas.reading import Reading, ReadingCreate, ReadingUpdate

router = APIRouter()

@router.get("/", response_model=List[Reading])
def list_readings(
    db: Session = Depends(get_db),
    q: Optional[str] = Query(None),
    category_id: Optional[int] = Query(None),
):
    qq = db.query(ReadingModel)
    if category_id is not None:
        qq = qq.filter(ReadingModel.category_id == category_id)
    if q:
        like = f"%{q}%"
        qq = qq.filter(ReadingModel.title.ilike(like))
    return qq.order_by(ReadingModel.id.desc()).all()

@router.post("/", response_model=Reading)
def create_reading(
    reading_in: ReadingCreate,
    db: Session = Depends(get_db),
    admin=Depends(get_current_staff_user),
):
    rd = ReadingModel(**reading_in.model_dump())
    db.add(rd)
    db.commit()
    db.refresh(rd)
    return rd

# ----------------- Bulk Import & Template -----------------
@router.get("/import-template.xlsx")
def download_template():
    try:
        import openpyxl
    except Exception as e:
        raise HTTPException(status_code=500, detail="openpyxl not installed on server")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "readings"
    headers = ["title", "content_html"]
    ws.append(headers)
    # Example rows
    ws.append(["Sample Reading A", "<p>Passage A text...</p>"])
    ws.append(["Sample Reading B", "<p>Passage B text...</p>"])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=readings-import-template.xlsx"},
    )


@router.post("/import-xlsx")
def import_readings(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    admin=Depends(get_current_staff_user),
):
    try:
        import openpyxl
    except Exception:
        raise HTTPException(status_code=500, detail="openpyxl not installed on server")
    try:
        # Read all bytes then load from a fresh BytesIO buffer (avoids pointer/seek issues)
        raw = file.file.read()
        if not raw:
            raise ValueError("empty")
        wb = openpyxl.load_workbook(filename=BytesIO(raw))
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid XLSX file")
    ws = wb.active
    # Map headers
    header_row = [str(c.value or '').strip() for c in next(ws.iter_rows(min_row=1, max_row=1))[0:]]
    # Normalize
    header_map = {h.lower().strip(): i for i, h in enumerate(header_row)}
    required = ["title", "content_html"]
    for h in required:
        if h not in header_map:
            raise HTTPException(status_code=400, detail=f"Missing column: {h}")
    created = 0
    errors: List[str] = []
    for r_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        def get(col):
            i = header_map.get(col)
            if i is None:
                return None
            val = row[i].value
            return str(val) if val is not None else ''
        title = (get("title") or '').strip()
        content = get("content_html") or ''
        if not title:
            continue
        try:
            rd = ReadingModel(title=title, content_html=content or '')
            db.add(rd)
            db.commit()
            db.refresh(rd)
            created += 1
        except Exception as e:
            db.rollback()
            errors.append(f"Row {r_idx}: {str(e)}")
    return {"created": created, "errors": errors}

# ID-based endpoints declared AFTER static routes to avoid capture of static paths
@router.get("/{reading_id}", response_model=Reading)
def get_reading(
    reading_id: int,
    db: Session = Depends(get_db),
    admin=Depends(get_current_staff_user),
):
    rd = db.query(ReadingModel).filter(ReadingModel.id == reading_id).first()
    if not rd:
        raise HTTPException(status_code=404, detail="Reading not found")
    return rd

@router.put("/{reading_id}", response_model=Reading)
def update_reading(
    reading_id: int,
    reading_in: ReadingUpdate,
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin_user),
):
    rd = db.query(ReadingModel).filter(ReadingModel.id == reading_id).first()
    if not rd:
        raise HTTPException(status_code=404, detail="Reading not found")
    for k, v in reading_in.model_dump(exclude_unset=True).items():
        setattr(rd, k, v)
    db.commit()
    db.refresh(rd)
    return rd

@router.delete("/{reading_id}")
def delete_reading(
    reading_id: int, db: Session = Depends(get_db), admin=Depends(get_current_admin_user)
):
    rd = db.query(ReadingModel).filter(ReadingModel.id == reading_id).first()
    if not rd:
        raise HTTPException(status_code=404, detail="Reading not found")
    db.delete(rd)
    db.commit()
    return {"message": "Reading deleted"}
